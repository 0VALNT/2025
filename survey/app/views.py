import csv

from django.shortcuts import redirect
from rest_framework import status, generics
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.generics import UpdateAPIView, ListCreateAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework import viewsets
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from django_filters import rest_framework as filters
from django.core.mail import send_mail
from .filters import SurveyFilter
from .models import Survey, Answer, SomeModel, AdminSurveyData, AdminStrQuestionSurveyData, \
    AdminChoseQuestionSurveyData, AdminIntQuestionSurveyData, CountOfAnswers, UserAnswers, UserAnswer
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.views import APIView
from . import serializers
from django.http import HttpResponse
from django.contrib.auth.models import User
from rest_framework_simplejwt.views import TokenObtainPairView
import openpyxl
from openpyxl.styles import Font


def login(request):
    token = request.COOKIES['access']
    if token:
        try:
            validated_token = JWTAuthentication().get_validated_token(token)
            user = JWTAuthentication().get_user(validated_token)
            request.user = user
        except Exception as e:
            res = [True, request]
            return res
        res = [False, request]
        return res
    else:
        res = [True, request]
        return res


class Login(TokenObtainPairView):
    def post(self, request: Request, *args, **kwargs) -> Response:
        serializer = self.get_serializer(data=request.data)

        try:
            serializer.is_valid(raise_exception=True)
        except TokenError as e:
            raise InvalidToken(e.args[0])
        response = Response(serializer.validated_data, status=status.HTTP_200_OK)
        print(serializer.validated_data['access'])
        response.set_cookie('access', value=serializer.validated_data['access'])
        return response


class ProtectedViews(APIView):

    def get(self, request, token=None):
        if token:
            try:
                validated_token = JWTAuthentication().get_validated_token(token)
                user = JWTAuthentication().get_user(validated_token)
                request.user = user
            except Exception as e:
                raise AuthenticationFailed('Invalid token')

        return Response({"message": "This is a protected view!", "user": str(request.user)})


class SeeMyForm(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'form.html'

    def get(self, request, pk):
        log = login(request)
        if log[0]:
            return redirect('token_obtain_pair')
        else:
            request = log[1]
        try:
            survey = Survey.objects.get(id=pk)
        except:
            return 0
        if request.user in survey.past_users.all() and survey.one_attempt:
            return HttpResponse('Ты уже прошел эту форму')
        return Response({'survey': survey})

    def post(self, request, pk):
        log = login(request)
        if log[0]:
            return redirect('token_obtain_pair')
        else:
            request = log[1]
        try:
            survey = Survey.objects.get(id=pk)
        except:

            return 0
        if request.user in survey.past_users.all() and survey.one_attempt:
            return HttpResponse('Ты уже прошел эту форму')
        user = request.user
        serializer = serializers.UserAnswersSerializer(data={
            'user': user.id,
            'survey': survey.id,
        })
        if serializer.is_valid():
            obj = serializer.save()
            obj.save()
        else:
            return 0
        post = request.POST
        data = AdminSurveyData.objects.get(survey_id=survey.id)
        data.num_of_pass += 1
        data.count_of_question = len(survey.questions.all())
        data.save()
        for question in survey.questions.all():
            if question.type == 'choices':
                answer = Answer.objects.get(id=int(request.POST[f'{question.id}']))
                new = False
                try:
                    model = CountOfAnswers.objects.get(answer_id=answer.id)
                    model.count_of_answer += 1
                except:
                    model = CountOfAnswers(answer_id=answer.id, count_of_answer=1)
                    new = True
                model.save()
                try:
                    question_data = AdminChoseQuestionSurveyData.objects.get(question_id=question.id)
                    question_data.count += 1
                    question_data.save()
                    if new:
                        question_data.help_model.add(model.id)
                        question_data.save()
                except:
                    question_data = AdminChoseQuestionSurveyData(question=question, admin_survey=data, count=1)
                    question_data.save()
                    question_data.help_model.add(model.id)
                    question_data.save()

                answer = answer.title
            elif question.type == 'multiple_choice':
                answers = post.getlist(f'{question.id}')
                text_answer = ''
                try:
                    question_data = AdminChoseQuestionSurveyData.objects.get(question_id=question.id)
                    question_data.count += 1
                    question_data.save()
                except:
                    question_data = AdminChoseQuestionSurveyData(question=question, admin_survey=data, count=1)
                    question_data.save()
                for i in answers:
                    answer = Answer.objects.get(id=int(i))
                    new = False
                    try:
                        model = CountOfAnswers.objects.get(answer_id=answer.id)
                        model.count_of_answer += 1
                    except:
                        model = CountOfAnswers(answer_id=answer.id, count_of_answer=1)
                        new = True
                    model.save()
                    if new:
                        question_data.help_model.add(model.id)
                        question_data.save()
                    if len(question_data.help_model.all()) == 0:
                        question_data.help_model.add(model.id)
                        question_data.save()
                    text_answer += Answer.objects.get(id=int(i)).title + ' '
                answer = text_answer
            else:
                answer = request.POST[f'{question.id}']
                if question.type == 'int':
                    try:
                        question_data = AdminIntQuestionSurveyData.objects.get(question_id=question.id)
                    except:
                        question_data = AdminIntQuestionSurveyData(question_id=question.id, admin_survey=data)
                    question_data.count += 1
                    question_data.amount += int(answer)
                    question_data.save()
                elif question.type == 'string':
                    try:
                        question_data = AdminStrQuestionSurveyData.objects.get(question_id=question.id)
                    except:
                        question_data = AdminStrQuestionSurveyData(question_id=question.id, admin_survey=data)
                    question_data.count += 1
                    question_data.amount += len(answer)
                    question_data.save()
            serializer = serializers.UserAnswerSerializer(data={
                'answer': answer,
                'question': question.id,
                'user_answers': obj.id,
                'type': question.type
            })
            if serializer.is_valid():
                objec = serializer.save()
                objec.save()
        survey.past_users.add(user)
        send_mail(
            "Прохождение опроса",
            f"Вы успешно прошли опрос!\n С уважением команда Опросы.ru.",
            "098ertyo@gmail.com",
            [request.user.email],
            fail_silently=False,
        )
        return HttpResponse('Вы успешно прошли опрос')


class Profile(APIView):
    def get(self, request):
        log = login(request)
        if log[0]:
            return redirect('token_obtain_pair')
        else:
            request = log[1]
        serializer = serializers.UserSerializer(request.user, context={'request': request})
        return Response(serializer.data)

    def put(self, request, token):
        log = login(request)
        if log[0]:
            return redirect('token_obtain_pair')
        else:
            request = log[1]
        serializer = serializers.UserSerializer(data=request.data, instance=request.user)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({"post": serializer.data})


class SomeView(APIView):
    def get(self, request, pk):
        some = SomeModel.objects.get(id=pk)
        serializer = serializers.SomeModelSerializer(some)
        return Response(serializer.data)

    def put(self, request, pk):
        some = SomeModel.objects.get(id=pk)
        serializer = serializers.SomeModelSerializer(data=request.data, instance=some)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({"post": serializer.data})


class SomeList(ListCreateAPIView):
    queryset = SomeModel.objects.all()
    serializer_class = serializers.SomeModelSerializer
    filter_fields = (
        'email',
        'username',
    )


class ChangePasswordView(UpdateAPIView):
    serializer_class = serializers.ChangePasswordSerializer
    model = User

    def get_object(self, queryset=None):
        log = login(self.request)
        if log[0]:
            return redirect('token_obtain_pair')
        else:
            request = log[1]
        obj = self.request.user
        return obj

    def update(self, request, *args, **kwargs):
        self.object = self.get_object()
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            if not self.object.check_password(serializer.data.get("old_password")):
                return Response({"old_password": ["Wrong password."]}, status=status.HTTP_400_BAD_REQUEST)
            self.object.set_password(serializer.data.get("new_password"))
            self.object.save()
            response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'Password updated successfully',
                'data': []
            }
            return Response(response)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SurveyViewSet(viewsets.ModelViewSet):
    queryset = Survey.objects.all()
    serializer_class = serializers.SurveySerializer
    filter_backends = [filters.DjangoFilterBackend]
    filterset_class = SurveyFilter


class GenerateExcelView(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Survey Data'

        header_font = Font(bold=True)
        header = [
            "Survey ID",
            "Number of Passes",
            "Count of Questions",
            "Question ID",
            "Question Type",
            "Avg Amount/Length",
            "Answer Breakdown"
        ]

        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font

        row_num = 2
        for survey_data in AdminSurveyData.objects.all():
            for int_question_data in AdminIntQuestionSurveyData.objects.filter(admin_survey=survey_data):
                ws.cell(row=row_num, column=1, value=survey_data.id)
                ws.cell(row=row_num, column=2, value=survey_data.num_of_pass)
                ws.cell(row=row_num, column=3, value=survey_data.count_of_question)
                ws.cell(row=row_num, column=4, value=int_question_data.question.id)
                ws.cell(row=row_num, column=5, value="Integer")
                avg_value = str(
                    int_question_data.amount / int_question_data.count if int_question_data.count > 0 else '-')
                ws.cell(row=row_num, column=6, value=avg_value if avg_value != '-' else '-')

                row_num += 1

            for str_question_data in AdminStrQuestionSurveyData.objects.filter(admin_survey=survey_data):
                ws.cell(row=row_num, column=1, value=survey_data.id)
                ws.cell(row=row_num, column=2, value=survey_data.num_of_pass)
                ws.cell(row=row_num, column=3, value=survey_data.count_of_question)
                ws.cell(row=row_num, column=4, value=str_question_data.question.id)
                ws.cell(row=row_num, column=5, value="String")
                avg_value = str(
                    str_question_data.amount / str_question_data.count if str_question_data.count > 0 else '-')
                ws.cell(row=row_num, column=6, value=avg_value if avg_value != '-' else '-')

                row_num += 1

            for chose_question_data in AdminChoseQuestionSurveyData.objects.filter(admin_survey=survey_data):
                ws.cell(row=row_num, column=1, value=survey_data.id)
                ws.cell(row=row_num, column=2, value=survey_data.num_of_pass)
                ws.cell(row=row_num, column=3, value=survey_data.count_of_question)
                ws.cell(row=row_num, column=4, value=chose_question_data.question.id)
                ws.cell(row=row_num, column=5, value="Choice")
                ws.cell(row=row_num, column=6, value="-")
                ws.cell(row=row_num, column=7, value=str(chose_question_data))
                row_num += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=survey_data.xlsx'
        wb.save(response)

        return response


class GenerateCsvView(APIView):
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="survey_data.csv"'

        writer = csv.writer(response)

        header = [
            "Survey ID",
            "Number of Passes",
            "Count of Questions",
            "Question ID",
            "Question Type",
            "Avg Amount/Length",
            "Answer Breakdown"
        ]
        writer.writerow(header)

        for survey_data in AdminSurveyData.objects.all():
            for int_question_data in AdminIntQuestionSurveyData.objects.filter(admin_survey=survey_data):
                avg_value = str(
                    int_question_data.amount / int_question_data.count if int_question_data.count > 0 else '-')
                writer.writerow([
                    survey_data.id,
                    survey_data.num_of_pass,
                    survey_data.count_of_question,
                    int_question_data.question.id,
                    "Integer",
                    avg_value if avg_value != '-' else '-',
                    "-"
                ])

            for str_question_data in AdminStrQuestionSurveyData.objects.filter(admin_survey=survey_data):
                avg_value = str(
                    str_question_data.amount / str_question_data.count if str_question_data.count > 0 else '-')
                writer.writerow([
                    survey_data.id,
                    survey_data.num_of_pass,
                    survey_data.count_of_question,
                    str_question_data.question.id,
                    "String",
                    avg_value if avg_value != '-' else '-',
                    "-"
                ])

            for chose_question_data in AdminChoseQuestionSurveyData.objects.filter(admin_survey=survey_data):
                writer.writerow([
                    survey_data.id,
                    survey_data.num_of_pass,
                    survey_data.count_of_question,
                    chose_question_data.question.id,
                    "Choice",
                    "-",
                    str(chose_question_data)

                ])

        return response


class GenerateExcelUserAnswersView(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'User Answers Data'

        header_font = Font(bold=True)
        header = [
            "User Answers ID",
            "User ID",
            "Survey ID",
            "Question ID",
            "Answer",
            "Type"
        ]

        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font

        row_num = 2
        for user_answers in UserAnswers.objects.all():
            for user_answer in UserAnswer.objects.filter(user_answers=user_answers):
                ws.cell(row=row_num, column=1, value=user_answers.id)
                ws.cell(row=row_num, column=2, value=user_answers.user.id)
                ws.cell(row=row_num, column=3, value=user_answers.survey.id)
                ws.cell(row=row_num, column=4, value=user_answer.question.id)
                ws.cell(row=row_num, column=5, value=user_answer.answer)
                ws.cell(row=row_num, column=6, value=user_answer.type)

                row_num += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=user_answers_data.xlsx'
        wb.save(response)

        return response


class GenerateCsvUserAnswersView(APIView):
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="user_answers_data.csv"'

        writer = csv.writer(response)

        header = [
            "User Answers ID",
            "User ID",
            "Survey ID",
            "Question ID",
            "Answer",
            "Type"
        ]
        writer.writerow(header)

        for user_answers in UserAnswers.objects.all():
            for user_answer in UserAnswer.objects.filter(user_answers=user_answers):
                writer.writerow([
                    user_answers.id,
                    user_answers.user.id,
                    user_answers.survey.id,
                    user_answer.question.id,
                    user_answer.answer,
                    user_answer.type
                ])

        return response
